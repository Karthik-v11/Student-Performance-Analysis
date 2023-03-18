#################################################################
# STUDENT PERFORMANCE ANALYSIS
# Made by
# Karthik V, Balakrishna K, Mohith G R
#################################################################

# libraries used
from tkinter import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl as xl
from matplotlib.figure import Figure
import matplotlib.patches as mpatches
import os
import re
from functools import partial

# retreiving excel files from the program directory
filelist = []
for x in os.listdir():
    if x.endswith(".xlsx"):
        filelist.append(x)
batch = []
print(filelist)
for i in filelist:
    batch.append(re.findall(r'\d{4}', i))
batches = sum(batch, [])
print(batches)

# function to select batch and semester
def ok():
    global batch_selected
    batch_selected = variable.get()
    variable.set(batch_selected)
    global sem_selected
    sem_selected = variable2.get()
    variable2.set(sem_selected)

# function to count pass and fail students
def functionality(sheet):
    pass_count = 0
    back_count = 0
    pass_list = []
    print(sheet.max_row)
    for r in range(1, sheet.max_row):
        if sheet.cell(row=r, column=1).value == 1:
            row_start = r
            print(row_start)
            break
    for r in range(row_start, sheet.max_row + 1):
        for c in range(1, sheet.max_column):
            cell1 = sheet.cell(row=r, column=c)
            if cell1.value is None:
                continue
            else:
                if sheet.cell(row=r, column=2).value in fail_list:
                    pass_count -= 1
                    break
                if cell1.value == "F":
                    fail_list.append(sheet.cell(row=r, column=2).value)
                    back_count += 1
                    pass_count -= 1
                    break
        pass_list.append(sheet.cell(row=r, column=2).value)
        pass_count += 1
    count_dict['students_without_backlog'] = pass_count
    count_dict['students_with_backlog'] = back_count
    print(pass_count)
    print(pass_list)
    print(fail_list)
    print(count_dict)
    print(len(pass_list))

# function to display the academic performance graph
def graph():
    try:
        ok()
        sheets = []
        global batch_selected
        for i in filelist:
            whatfile = re.findall(r'\d{4}', i)
            if batch_selected == whatfile[0]:
                workbook = xl.load_workbook(i, data_only=True)
                sheets = workbook.sheetnames
                n = 1
                for i in sheets:
                    globals()[f"sheet{n}"] = workbook[i]
                    n += 1
    except:
        print("file format error")
    print(sem_selected)
    if sem_selected == "1st Sem":
        fail_list.clear()
        functionality(sheet1)
    elif sem_selected == "2nd Sem" or sem_selected == "1st Year":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
    elif sem_selected == "3rd Sem":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
    elif sem_selected == "4th Sem" or sem_selected == "2nd Year":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
        functionality(sheet4)
    elif sem_selected == "5th Sem":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
        functionality(sheet4)
        functionality(sheet5)
    elif sem_selected == "6th Sem" or sem_selected == "3rd Year":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
        functionality(sheet4)
        functionality(sheet5)
        functionality(sheet6)
    elif sem_selected == "7th Sem":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
        functionality(sheet4)
        functionality(sheet5)
        functionality(sheet6)
        functionality(sheet7)
    elif sem_selected == "8th Sem" or sem_selected == "4th Year":
        fail_list.clear()
        functionality(sheet1)
        functionality(sheet2)
        functionality(sheet3)
        functionality(sheet4)
        functionality(sheet5)
        functionality(sheet6)
        functionality(sheet7)
        functionality(sheet8)
    destroy()
    create()

# function to display the result graph
def resgraph():
    try:
        ok()
        sheets = []
        global batch_selected
        for i in filelist:
            whatfile = re.findall(r'\d{4}', i)
            if batch_selected == whatfile[0]:
                workbook = xl.load_workbook(i, data_only=True)
                sheets = workbook.sheetnames
                print(sheets)
                n = 1
                for i in sheets:
                    globals()[f"sheet{n}"] = workbook[i]
                    n += 1
    except:
        print("file format error")

    print(sem_selected)
    if sem_selected == "1st Sem":
        calculate_res(sheet1)
    elif sem_selected == "2nd Sem":
        calculate_res(sheet2)
    elif sem_selected == "1st Year":
        cal_year1_res()
    elif sem_selected == "3rd Sem":
        calculate_res(sheet3)
    elif sem_selected == "4th Sem":
        calculate_res(sheet4)
    elif sem_selected == "2nd Year":
        cal_year2_res()
    elif sem_selected == "5th Sem":
        calculate_res(sheet5)
    elif sem_selected == "6th Sem":
        calculate_res(sheet6)
    elif sem_selected == "3rd Year":
        cal_year3_res()
    elif sem_selected == "7th Sem":
        calculate_res(sheet7)
    elif sem_selected == "8th Sem":
        calculate_res(sheet8)
    elif sem_selected == "4th Year":
        cal_year4_res()
    destroy()
    createres()

# calculate the result for 1st year; 1sem and 2sem
def cal_year1_res():
    global agg
    agg = {}
    fcdcount = 0
    fccount = 0
    sccount = 0

    row_start1, col1 = row_col_start(sheet1)
    row_start2, col2 = row_col_start(sheet2)
    for r1 in range(row_start1, sheet1.max_row + 1):
        result_cell1 = sheet1.cell(row=r1, column=col1)
        if result_cell1.value is None or type(result_cell1.value) == str:
            continue
        else:
            for r2 in range(row_start2, sheet2.max_row + 1):
                result_cell2 = sheet2.cell(row=r2, column=col2)
                if result_cell2.value is None or type(result_cell2.value) == str:
                    continue
                else:
                    if (sheet1.cell(row=r1, column=2).value == sheet2.cell(row=r2, column=2).value):
                        agg_per = (sheet1.cell(row=r1, column=col1).value + sheet2.cell(row=r2, column=col2).value) / 2;
                        agg[sheet1.cell(row=r1, column=2).value] = agg_per
    print(agg)
    res_list = agg.values()
    for i in res_list:
        if i >= 70:
            fcdcount += 1
        elif i >= 60:
            fccount += 1
        elif i >= 35:
            sccount += 1
    result_dict['FCD'] = fcdcount
    result_dict['FC'] = fccount
    result_dict['SC'] = sccount

# calculate the result for 2nd year; 1st year with 3rd sem and 4th sem
def cal_year2_res():
    global agg
    fcdcount = 0
    fccount = 0
    sccount = 0
    cal_year1_res()
    row_start3, col3 = row_col_start(sheet3)
    row_start4, col4 = row_col_start(sheet4)
    for r3 in range(row_start3, sheet3.max_row + 1):
        result_cell3 = sheet3.cell(row=r3, column=col3)
        if result_cell3.value is None or type(result_cell3.value) == str:
            continue
        else:
            for r4 in range(row_start4, sheet4.max_row + 1):
                result_cell4 = sheet4.cell(row=r4, column=col4)
                if result_cell4.value is None or type(result_cell4.value) == str:
                    continue
                else:
                    if (sheet3.cell(row=r3, column=2).value == sheet4.cell(row=r4, column=2).value):
                        agg_per = (sheet3.cell(row=r3, column=col3).value + sheet4.cell(row=r4, column=col4).value) / 2;
                        agg[sheet1.cell(row=r3, column=2).value] += agg_per
                        agg[sheet1.cell(row=r3, column=2).value] /= 2
    res_list = agg.values()
    for i in res_list:
        if i >= 70:
            fcdcount += 1
        elif i >= 60:
            fccount += 1
        elif i >= 35:
            sccount += 1
    result_dict['FCD'] = fcdcount
    result_dict['FC'] = fccount
    result_dict['SC'] = sccount

# calculate the result for 3rd year; 1st year and 2nd year with 5th sem and 6th sem
def cal_year3_res():
    global agg
    fcdcount = 0
    fccount = 0
    sccount = 0
    cal_year1_res()
    cal_year2_res()
    row_start5, col5 = row_col_start(sheet5)
    row_start6, col6 = row_col_start(sheet6)
    for r5 in range(row_start5, sheet5.max_row + 1):
        result_cell5 = sheet5.cell(row=r5, column=col5)
        if result_cell5.value is None or type(result_cell5.value) == str:
            continue
        else:
            for r6 in range(row_start6, sheet6.max_row + 1):
                result_cell6 = sheet6.cell(row=r6, column=col6)
                if result_cell6.value is None or type(result_cell6.value) == str:
                    continue
                else:
                    if (sheet5.cell(row=r5, column=2).value == sheet6.cell(row=r6, column=2).value):
                        agg_per = (sheet5.cell(row=r5, column=col5).value + sheet6.cell(row=r6, column=col6).value) / 2;
                        agg[sheet1.cell(row=r5, column=2).value] += agg_per
                        agg[sheet1.cell(row=r5, column=2).value] /= 2
    res_list = agg.values()
    for i in res_list:
        if i >= 70:
            fcdcount += 1
        elif i >= 60:
            fccount += 1
        elif i >= 35:
            sccount += 1
    result_dict['FCD'] = fcdcount
    result_dict['FC'] = fccount
    result_dict['SC'] = sccount

# calculate the result for 4th year; 1st,2nd and 3rd year with 7th sem and 8th sem
def cal_year4_res():
    global agg
    fcdcount = 0
    fccount = 0
    sccount = 0
    cal_year1_res()
    cal_year2_res()
    cal_year3_res()
    row_start7, col7 = row_col_start(sheet7)
    row_start8, col8 = row_col_start(sheet8)
    for r7 in range(row_start7, sheet7.max_row + 1):
        result_cell7 = sheet7.cell(row=r7, column=col7)
        if result_cell7.value is None or type(result_cell7.value) == str:
            continue
        else:
            for r8 in range(row_start8, sheet8.max_row + 1):
                result_cell8 = sheet8.cell(row=r8, column=col8)
                if result_cell8.value is None or type(result_cell8.value) == str:
                    continue
                else:
                    if (sheet7.cell(row=r7, column=2).value == sheet8.cell(row=r8, column=2).value):
                        agg_per = (sheet7.cell(row=r7, column=col7).value + sheet8.cell(row=r8, column=col8).value) / 2;
                        agg[sheet1.cell(row=r7, column=2).value] += agg_per
                        agg[sheet1.cell(row=r7, column=2).value] /= 2
    res_list = agg.values()
    for i in res_list:
        if i >= 70:
            fcdcount += 1
        elif i >= 60:
            fccount += 1
        elif i >= 35:
            sccount += 1
    result_dict['FCD'] = fcdcount
    result_dict['FC'] = fccount
    result_dict['SC'] = sccount

# function to choose row and column
def row_col_start(sheet):
    for r in range(1, sheet.max_row):
        if sheet.cell(row=r, column=1).value == 1:
            row_start = r
            break
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            cell = str(sheet.cell(row=r, column=c).value)
            if cell.__contains__('%'):
                col = c
                break
    return row_start, col

# function to find the subject result
def subject_res():
    try:
        ok()
        sheets = []
        global batch_selected
        for i in filelist:
            whatfile = re.findall(r'\d{4}', i)
            if batch_selected == whatfile[0]:
                workbook = xl.load_workbook(i, data_only=True)
                print(batch_selected)
                sheets = workbook.sheetnames
                print(sheets)
                n = 1
                for i in sheets:
                    globals()[f"sheet{n}"] = workbook[i]
                    n += 1
    except:
        print("file format error")
    if sem_selected == "1st Sem":
        subjects(sheet1)
    elif sem_selected == "2nd Sem":
        subjects(sheet2)
    elif sem_selected == "3rd Sem":
        subjects(sheet3)
    elif sem_selected == "4th Sem":
        subjects(sheet4)
    elif sem_selected == "5th Sem":
        subjects(sheet5)
    elif sem_selected == "6th Sem":
        subjects(sheet6)
    elif sem_selected == "7th Sem":
        subjects(sheet7)
    elif sem_selected == "8th Sem":
        subjects(sheet8)

# function to extract the subjects from the sheet
def subjects(sheet):
    subject = []
    for r in range(1, sheet.max_row):
        for c in range(1, sheet.max_column):
            if sheet.cell(row=r, column=c).value == None or type(sheet.cell(row=r, column=c).value) == int:
                continue
            else:
                pattern = r'(^\d{2})([a-zA-Z]{2,4})(\d{2})'
                for match in re.finditer(pattern, str(sheet.cell(row=r, column=c).value)):
                    x = match.group()
                    subject.append(x)

    subject_codes = list(set(subject))
    print(subject_codes)

    SUBJECTS = subject_codes
    global variable3
    variable3 = StringVar(root)
    variable3.set(SUBJECTS[0])
    w3 = OptionMenu(root, variable3, *SUBJECTS)
    w3.place(relx=0.8, rely=0.25, anchor='ne')

    sub_button = Button(root, text="OK", fg='black', bg='#b9d9eb', command=partial(okie, sheet))
    sub_button.pack()
    sub_button.place(relx=0.8, rely=0.32, anchor=CENTER)

# function to calculate the subject result
def sub_calculate(sheet):
    subpasscount = 0
    subfailcount = 0

    global sub_selected

    print(sub_selected)

    for r in range(1, sheet.max_row):
        for c in range(1, sheet.max_column):
            if sheet.cell(row=r, column=c).value == None or type(sheet.cell(row=r, column=c).value) == int:
                continue
            else:
                if sheet.cell(row=r, column=c).value == sub_selected:
                    for row in range(r, sheet.max_row):
                        if (sheet.cell(row=row + 2, column=c + 3).value == 'P'):
                            subpasscount += 1
                        elif (sheet.cell(row=row + 2, column=c + 3).value == 'F'):
                            subfailcount += 1

    subcount_dict['no of students passed'] = subpasscount
    subcount_dict['no of students failed'] = subfailcount

    destroy()
    subgraph()


def okie(sheet):
    global sub_selected
    sub_selected = variable3.get()
    variable3.set(sub_selected)
    sub_calculate(sheet)


def calculate_res(sheet):
    global result_dict
    fcdcount = 0
    fccount = 0
    sccount = 0
    sclist = []
    row_start, col = row_col_start(sheet)
    for r in range(row_start, sheet.max_row + 1):
        result_cell = sheet.cell(row=r, column=col)
        if result_cell.value is None or type(result_cell.value) == str:
            continue
        else:
            if result_cell.value >= 70:
                fcdcount += 1
            elif result_cell.value >= 60:
                fccount += 1
            elif result_cell.value >= 35:
                sclist.append(sheet.cell(row=r, column=sheet.max_column).value)
                sccount += 1
    result_dict['FCD'] = fcdcount
    result_dict['FC'] = fccount
    result_dict['SC'] = sccount
    print(result_dict)
    print(sclist)


def subgraph():
    global fig
    global canvas
    colors = ['green', 'red']
    no_backlogs = mpatches.Patch(color='green', label='Pass')
    yes_backlogs = mpatches.Patch(color='red', label='Fail')
    fig = Figure(figsize=(4.5, 4.5), dpi=100)
    plot1 = fig.add_subplot(111)
    plot1.bar(subcount_dict.keys(), subcount_dict.values(), width=0.4, color=colors)
    plot1.set_ylabel('number of students')
    plot1.set_xlabel(batch_selected + " batch " + sem_selected + " results")
    key1 = subcount_dict['no of students passed']
    plot1.text(0, 1, key1)
    key2 = subcount_dict['no of students failed']
    plot1.text(1, 1, key2)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=BOTTOM)
    fig.legend(handles=[no_backlogs, yes_backlogs])


def createres():
    global fig
    global canvas
    colors = ['green', 'red', 'blue']
    fcd = mpatches.Patch(color='green', label='FCD')
    fc = mpatches.Patch(color='red', label='FC')
    sc = mpatches.Patch(color='blue', label='SC')
    fig = Figure(figsize=(4, 4), dpi=105)
    plot1 = fig.add_subplot(111)
    plot1.bar(result_dict.keys(), result_dict.values(), width=0.4, color=colors)
    plot1.set_ylabel('number of students')
    plot1.set_xlabel(batch_selected + " batch " + sem_selected + " results")
    key1 = result_dict['FCD']
    plot1.text(0, 1, key1)
    key2 = result_dict['FC']
    plot1.text(1, 1, key2)
    key3 = result_dict['SC']
    plot1.text(2, 1, key3)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=BOTTOM)
    fig.legend(handles=[fcd, fc, sc])


def create():
    global fig
    global canvas
    colors = ['green', 'red']
    no_backlogs = mpatches.Patch(color='green', label='students_without_backlogs')
    yes_backlogs = mpatches.Patch(color='red', label='students_with_backlogs')
    fig = Figure(figsize=(4.5, 4.5), dpi=100)
    plot1 = fig.add_subplot(111)
    plot1.bar(count_dict.keys(), count_dict.values(), width=0.4, color=colors)
    plot1.set_ylabel('number of students')
    plot1.set_xlabel(batch_selected + " batch " + sem_selected + " results")
    key1 = count_dict['students_without_backlog']
    plot1.text(0, 1, key1)
    key2 = count_dict['students_with_backlog']
    plot1.text(1, 1, key2)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(side=BOTTOM)
    fig.legend(handles=[no_backlogs, yes_backlogs])


def destroy():
    global canvas
    canvas.get_tk_widget().destroy()


def save():
    global fig
    print('saving...')
    fig.savefig(batch_selected + " " + sem_selected + '.png', dpi=500)


# the main display of the application
root = Tk()
root.title('Student performance analysis')
root.geometry("1280x720")
frame = Frame(root, width=50, height=40)
frame.pack()

bg = PhotoImage(file="Background.png")

label2 = Label(root, image=bg)
label2.place(x=0, y=0)

OPTIONS = batches

variable = StringVar(root)
variable.set(OPTIONS[0])
w = OptionMenu(root, variable, *OPTIONS)
w.place(relx=0.5, rely=0.25, anchor='ne')

OPTIONS2 = [
    "1st Sem",
    "2nd Sem",
    "1st Year",
    "3rd Sem",
    "4th Sem",
    "2nd Year",
    "5th Sem",
    "6th Sem",
    "3rd Year",
    "7th Sem",
    "8th Sem",
    "4th Year"
]

variable2 = StringVar(root)
variable2.set(OPTIONS2[0])
w2 = OptionMenu(root, variable2, *OPTIONS2)
w2.place(relx=0.5, rely=0.25, anchor="nw")

batch_selected = ""
sub_selected = ""
sem_selected = ""
fail_list = []
count_dict = {}
result_dict = {}
subcount_dict = {}

canvas = FigureCanvasTkAgg(master=root)
fig = Figure(figsize=(4, 4), dpi=100)

perf_button = Button(root, text="Academic Performance", fg='black', bg='#b9d9eb', command=graph)
perf_button.pack()
perf_button.place(relx=0.40, rely=0.32, anchor=CENTER)

rst_button = Button(root, text="Final Result", fg='black', bg='#b9d9eb', command=resgraph)
rst_button.pack()
rst_button.place(relx=0.50, rely=0.32, anchor=CENTER)

rst_button = Button(root, text="Subject Result", fg='black', bg='#b9d9eb', command=subject_res)
rst_button.pack()
rst_button.place(relx=0.60, rely=0.32,anchor=CENTER)

save_button = Button(root, text="Save", fg='black', bg='#b9d9eb', command=save)
save_button.pack(side=BOTTOM)

root.mainloop()
